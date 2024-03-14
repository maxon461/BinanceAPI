from openpyxl import Workbook
import locale

def save_to_excel(buy_orders, sell_orders, profit, filename):
    # Set the locale to format numbers with commas
    locale.setlocale(locale.LC_NUMERIC, 'en_US.UTF-8')

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write headers for buy orders
    headers = ["Order Status", "Total Price", "Trade Type", "Create Time", "Profit"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write buy orders data starting from cell A2
    for i, order in enumerate(buy_orders, start=2):
        total_price = int(locale.atof(order["totalPrice"]))
        row = [
            order["orderStatus"],
            total_price,
            order["tradeType"],
            order["createTime"],
            ""  # Placeholder for profit, will be filled later
        ]
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    # Write headers for sell orders starting from cell H1
    sell_headers = ["Order Status", "Total Price", "Trade Type", "Create Time", "Profit"]
    for col, header in enumerate(sell_headers, start=1):
        ws.cell(row=1, column=col+7, value=header)

    # Write sell orders data starting from cell H2
    for i, order in enumerate(sell_orders, start=2):
        total_price = int(locale.atof(order["totalPrice"]))
        row = [
            order["orderStatus"],
            total_price,
            order["tradeType"],
            order["createTime"],
            ""  # Placeholder for profit, will be filled later
        ]
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j+7, value=value)

    # Write profit for sell orders
    ws.cell(row=1, column=13, value="Profit")
    ws.cell(row=2, column=13, value=profit)

    # Save workbook to Excel file
    wb.save(filename)